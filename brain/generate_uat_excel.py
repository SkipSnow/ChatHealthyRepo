import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Load JSON data
with open('brain/assurance_results.json', 'r') as file:
    data = json.load(file)

# Find the assignment entry
assignment_id = 'ASN-1F9268'
raw = data[assignment_id]["content"]
import re
raw = re.sub(r'^```json\s*', '', raw.strip())
raw = re.sub(r'\s*```$', '', raw.strip())
assignment_data = json.loads(raw)

# Extract test suite
excel_test_suite = assignment_data['excel_test_suite']

# Create a new workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True)
true_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
false_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Add sheets and populate data
for feature in excel_test_suite:
    ws = wb.create_sheet(title=feature['feature'][:31].replace('/', '-'))
    ws.append([feature['feature']])
    ws.append([feature['description']])
    ws.append(['Test ID', 'Scenario', 'Acceptance Criterion', 'Expected Result'])
    for cell in ws[3]:
        cell.font = header_font
    for index, test_case in enumerate(feature['test_cases'], start=4):
        ws.append([
            test_case['test_id'],
            test_case['scenario'],
            test_case['acceptance_criterion'],
            'TRUE' if test_case['expected_result'] else 'FALSE'
        ])
        # Apply fill color based on expected result
        fill = true_fill if test_case['expected_result'] else false_fill
        ws[f'D{index}'].fill = fill
        # Alternate row colors for readability
        if index % 2 == 0:
            for cell in ws[index]:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# Remove default sheet
wb.remove(wb['Sheet'])

# Save the workbook
wb.save('uat_test_cases.xlsx')
