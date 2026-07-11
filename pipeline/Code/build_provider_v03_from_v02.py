"""build_provider_v03_from_v02.py

Production transform: reads PublicHealthData.provider_v02, writes
PublicHealthData.provider_v03 on the front-end MongoDB cluster.

Run from repo root:
    python pipeline/Code/build_provider_v03_from_v02.py
    python pipeline/Code/build_provider_v03_from_v02.py --limit 100
    python pipeline/Code/build_provider_v03_from_v02.py --drop

Transformation contract
-----------------------
For every v02 document the output v03 document is:
    v02 document
    - optionally augmented with `licenses` (existing + state licenses
      mined from `other_identifiers` whose issuer names a state board
      / state department of health / explicit license token)
    - optionally augmented with `insurance` (Medicaid, Medicare UPIN,
      commercial issuer-driven entries, and unclassified-issuer fall-
      through entries mined from `other_identifiers`)
    - the v02 _id is stripped (Mongo assigns a fresh _id on insert).

Documents with `out_of_scope.flagged == True` are NOT written to v03.

v02 is never mutated. v03 is the only write target.

OI classifier (per OI entry)
----------------------------
The classifier replaces the prior "any non-empty issuer → Commercial"
shortcut, which was wrong: many issuer strings name a state board, a
state department of health, or are literal license tokens. Those are
state licenses, not insurance, and routing them to insurance[]
pollutes carrier facets downstream.

Seeded by the two analysis workbooks:
    _oneshots/test_output/v02_OI_Issuer_Analysis.xlsx
    _oneshots/test_output/v02_OI_StateBoard_Analysis.xlsx

Per OI entry:
1. STATE LICENSE detection — match if ANY of the following are true
   (issuer matching uses the same uppercase+whitespace-collapse+
   trailing-punctuation-strip normalization as the workbooks):
     a. The normalized issuer is in the workbook's STATE_BOARD set
        (~4,451 distinct strings).
     b. The normalized issuer contains one of the canonical board
        / department-of-health phrases (see _STATE_BOARD_PHRASES).
     c. The normalized issuer is a literal license token
        ("LICENSE", "STATE LICENSE", "MEDICAL LICENSE",
        "MEDICINE LICENSE", "NURSING LICENSE", "PHARMACY LICENSE",
        "LICENSE NUMBER").
     d. The OI entry's type_description (upper-cased) contains the
        word "LICENSE".
   Matched entries emit to licenses[] as {state, number} — ONLY those
   two fields, matching the v02 license schema exactly.

2. INSURANCE detection (only if rule 1 did not fire):
     a. type_code in {"02","05"} or "MEDICAID" in type_description
        → Medicaid.
     b. type_code == "04" or "UPIN" in type_description
        → Medicare UPIN.
     c. Normalized issuer is in the workbook's COMMERCIAL_CARRIER
        set OR contains one of the canonical carrier phrases (see
        _COMMERCIAL_CARRIER_PHRASES) → Commercial.

3. AMBIGUOUS / UNRECOGNIZED fall-through:
     - issuer non-empty → emit to insurance[] with insurance_type=""
       (the honest "we couldn't classify" signal).
     - issuer empty AND no other rule fired → increment
       dropped_by_type_code; do not emit.

License merge semantics
-----------------------
Three-pass dedup; final licenses[] is guaranteed to contain no two
entries with the same dedup key.

  Pass 1: dedup v02's own licenses[] on (upper(state),
          normalized(number)). First occurrence wins. Silent. Counted
          in v02_license_dupes_dropped.
  Pass 2: when adding an OI-derived license, drop any candidate whose
          key is already in the deduped v02 set. Log a "license
          dispute" warning and increment license_disputes.
  Pass 3: dedup OI-derived licenses against each other on the same
          key. First wins. Silent. Counted in
          oi_license_dupes_dropped.

If licenses[] would be empty after processing, the field is omitted
from the v03 document entirely.

License shape — STRICT
----------------------
Each licenses[] entry is exactly:
    {"state": <USPS 2-letter, uppercase>, "number": <trimmed>}
No source annotation, no classification metadata, no issuer field.
v02 licenses pass through with whatever extra fields they already
carry; OI-derived licenses are minted with only the two fields above.

Insurance shape
---------------
Each insurance[] entry is a flat object:
    {"insurance_type": str, "brand": str, "state": str, "raw_value": str}
Dedup is by (upper(insurance_type), upper(brand).strip(),
upper(state).strip(), raw_value.strip()).
If insurance[] would be empty, the field is omitted.

`carriers[]` is gone. The v03 output never carries one.

Everything else passes through verbatim.

Regex policy (Rule-008 statement 4)
-----------------------------------
No `re` import. Issuer normalization and license normalization use
only plain string operations.
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
import openpyxl
from pymongo import MongoClient, ReplaceOne
from pymongo.errors import BulkWriteError

REPO_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(REPO_ROOT / "Code" / ".env")

DB_NAME = "PublicHealthData"
SRC_COLLECTION = "provider_v02"
DST_COLLECTION = "provider_v03"

DEFAULT_BATCH_SIZE = 1_000
PROGRESS_EVERY = 10_000

ISSUER_ANALYSIS_XLSX = (
    REPO_ROOT / "_oneshots" / "test_output" / "v02_OI_Issuer_Analysis.xlsx"
)

log = logging.getLogger("build_provider_v03_from_v02")


def _setup_logging() -> None:
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    root = logging.getLogger()
    root.handlers[:] = [handler]
    root.setLevel(logging.INFO)


# ---------------------------------------------------------------------------
# Issuer normalization (matches the workbook's normalization exactly)
# ---------------------------------------------------------------------------

_TRAILING_PUNCT = ",.;:- "


def _normalize_issuer(raw: Any) -> str:
    """Uppercase, collapse internal whitespace runs to single spaces,
    strip leading/trailing whitespace and trailing punctuation. No regex."""
    if raw is None:
        return ""
    s = str(raw).upper()
    # Collapse internal whitespace runs to single spaces (no regex).
    parts = s.split()
    s = " ".join(parts)
    # Strip leading/trailing punctuation runs.
    while s and s[-1] in _TRAILING_PUNCT:
        s = s[:-1]
    while s and s[0] in _TRAILING_PUNCT:
        s = s[1:]
    return s


# ---------------------------------------------------------------------------
# Classifier reference data — seeded from the issuer-analysis workbook.
# ---------------------------------------------------------------------------

# Canonical state-board / state-DOH phrases (uppercase). Substring match
# against the normalized issuer. Derived from the by_board_type tab of
# v02_OI_StateBoard_Analysis.xlsx and the proposed_classification
# rationales of v02_OI_Issuer_Analysis.xlsx.
_STATE_BOARD_PHRASES: tuple[str, ...] = (
    "BOARD OF MEDICINE",
    "MEDICAL BOARD",
    "BOARD OF NURSING",
    "NURSING BOARD",
    "BOARD OF PHARMACY",
    "PHARMACY BOARD",
    "BOARD OF DENTAL",
    "DENTAL EXAMINERS",
    "BOARD OF CHIROPRACTIC",
    "BOARD OF OPTOMETRY",
    "OPTOMETRIC BOARD",
    "BOARD OF PSYCHOLOGY",
    "BOARD OF SOCIAL WORK",
    "SOCIAL WORK BOARD",
    "BOARD OF COUNSELING",
    "MARRIAGE AND FAMILY",
    "BOARD OF HEALING ARTS",
    "BOARD OF PHYSICAL THERAPY",
    "PHYSICAL THERAPY EXAMINERS",
    "BOARD OF OSTEOPATHIC",
    "OSTEOPATHIC BOARD",
    "BOARD OF HEALTH PROFESSIONS",
    "DEPARTMENT OF HEALTH",
    "STATE BOARD",
    "STATE LICENSE",
    "MEDICAL LICENSE",
    "MEDICINE LICENSE",
    "NURSING LICENSE",
    "PHARMACY LICENSE",
    "LICENSE NUMBER",
)

# Literal-license tokens (the entire normalized issuer equals one of
# these). High-frequency UNRECOGNIZED entries that prior analysis
# identified as licenses-in-disguise.
_LITERAL_LICENSE_TOKENS: frozenset[str] = frozenset({
    "LICENSE",
    "STATE LICENSE",
    "MEDICAL LICENSE",
    "MEDICINE LICENSE",
    "NURSING LICENSE",
    "PHARMACY LICENSE",
    "LICENSE NUMBER",
})

# Canonical commercial-carrier phrases (uppercase). Substring match.
_COMMERCIAL_CARRIER_PHRASES: tuple[str, ...] = (
    "AETNA",
    "AETENA",  # observed misspelling
    "BLUE CROSS",
    "BLUE SHIELD",
    "BCBS",
    "BC/BS",
    "CIGNA",
    "HUMANA",
    "ANTHEM",
    "UHC",
    "UNITEDHEALTH",
    "KAISER",
    "MOLINA",
    "CENTENE",
    "WELLCARE",
    "MAGELLAN",
    "OPTUM",
    "EXPRESS SCRIPTS",
    "CAREMARK",
    "HEALTHLINK",
    "RAILROAD MEDICARE",
    "TRICARE",
    "CHAMPVA",
)

# Sets loaded from the workbook at startup. Populated by
# _load_issuer_sets(); empty until then.
_STATE_BOARD_SET: set[str] = set()
_COMMERCIAL_CARRIER_SET: set[str] = set()


def _load_issuer_sets(xlsx_path: Path = ISSUER_ANALYSIS_XLSX) -> None:
    """Read the proposed_classification tab and populate the
    STATE_BOARD and COMMERCIAL_CARRIER sets. Idempotent — safe to
    call from both the production main() and from the smoke wrapper."""
    global _STATE_BOARD_SET, _COMMERCIAL_CARRIER_SET
    if _STATE_BOARD_SET and _COMMERCIAL_CARRIER_SET:
        return
    if not xlsx_path.exists():
        log.warning("issuer analysis workbook not found at %s — "
                    "classifier will rely on phrase rules only",
                    xlsx_path)
        return
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "proposed_classification" not in wb.sheetnames:
            log.warning("proposed_classification tab missing in %s", xlsx_path)
            return
        ws = wb["proposed_classification"]
        sb: set[str] = set()
        cc: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            iss = row[0]
            cls = row[2] if len(row) >= 3 else None
            if iss is None or cls is None:
                continue
            iss_norm = _normalize_issuer(iss)
            if not iss_norm:
                continue
            if cls == "STATE_BOARD":
                sb.add(iss_norm)
            elif cls == "COMMERCIAL_CARRIER":
                cc.add(iss_norm)
        _STATE_BOARD_SET = sb
        _COMMERCIAL_CARRIER_SET = cc
        log.info("loaded issuer sets: STATE_BOARD=%d, COMMERCIAL_CARRIER=%d",
                 len(_STATE_BOARD_SET), len(_COMMERCIAL_CARRIER_SET))
    finally:
        wb.close()


def _issuer_is_state_license(issuer_norm: str, type_description_upper: str) -> bool:
    """True iff the OI entry is a state license per the new classifier."""
    if type_description_upper and "LICENSE" in type_description_upper:
        return True
    if not issuer_norm:
        return False
    if issuer_norm in _LITERAL_LICENSE_TOKENS:
        return True
    if issuer_norm in _STATE_BOARD_SET:
        return True
    for phrase in _STATE_BOARD_PHRASES:
        if phrase in issuer_norm:
            return True
    return False


def _issuer_is_commercial_carrier(issuer_norm: str) -> bool:
    """True iff the normalized issuer names a known commercial carrier."""
    if not issuer_norm:
        return False
    if issuer_norm in _COMMERCIAL_CARRIER_SET:
        return True
    for phrase in _COMMERCIAL_CARRIER_PHRASES:
        if phrase in issuer_norm:
            return True
    return False


# ---------------------------------------------------------------------------
# License number normalization (no regex)
# ---------------------------------------------------------------------------

def _normalize_state(state: Any) -> str:
    return str(state or "").strip().upper()


def _strip_separators_and_zeros(raw: str) -> str:
    kept_chars = []
    for ch in raw:
        if ch.isalnum():
            kept_chars.append(ch)
    s = "".join(kept_chars)
    s = s.lstrip("0")
    if not s:
        return "0"
    return s


def _license_dedup_key(state: Any, number: Any) -> tuple[str, str]:
    st = _normalize_state(state)
    num_str = str(number or "").upper()
    num_norm = _strip_separators_and_zeros(num_str)
    return st, num_norm


# ---------------------------------------------------------------------------
# Transform
# ---------------------------------------------------------------------------

class TransformCounters:
    def __init__(self) -> None:
        self.scanned = 0
        self.written = 0
        self.failed_docs = 0
        self.skipped_out_of_scope = 0
        self.dropped_by_reason: Counter = Counter()
        self.v02_license_dupes_dropped = 0
        self.oi_license_dupes_dropped = 0
        self.state_license_added = 0
        self.medicaid_insurance_added = 0
        self.upin_insurance_added = 0
        self.commercial_insurance_added = 0
        self.unclassified_insurance_added = 0
        self.license_disputes = 0
        self.dropped_by_type_code: Counter = Counter()


def transform_doc(doc: dict, counters: TransformCounters) -> dict | None:
    """Return the v03 document, or None if the doc is out-of-scope."""
    oos = doc.get("out_of_scope") or {}
    if isinstance(oos, dict) and oos.get("flagged") is True:
        counters.skipped_out_of_scope += 1
        reason = oos.get("reason") or "<missing>"
        counters.dropped_by_reason[str(reason)] += 1
        return None

    npi = doc.get("npi")

    # --- Pass 1: dedup v02's own licenses[]. ---
    raw_licenses = doc.get("licenses") or []
    deduped_v02_licenses: list[dict] = []
    seen_license_keys: set[tuple[str, str]] = set()
    for lic in raw_licenses:
        if not isinstance(lic, dict):
            continue
        st = lic.get("state")
        num = lic.get("number")
        if st and num:
            key = _license_dedup_key(st, num)
            if key in seen_license_keys:
                counters.v02_license_dupes_dropped += 1
                continue
            seen_license_keys.add(key)
        deduped_v02_licenses.append(lic)

    final_licenses = list(deduped_v02_licenses)
    insurance: list[dict] = []
    seen_insurance_keys: set[tuple[str, str, str, str]] = set()

    raw_oi = doc.get("other_identifiers") or []
    if not isinstance(raw_oi, list):
        raw_oi = []

    for ident in raw_oi:
        if not isinstance(ident, dict):
            continue
        raw_id = ident.get("identifier")
        if raw_id is None or str(raw_id).strip() == "":
            continue

        raw_id_str = str(raw_id).strip()
        tc_raw = ident.get("type_code")
        type_code = str(tc_raw or "").strip().zfill(2) if tc_raw is not None else ""
        desc_upper = str(ident.get("type_description") or "").upper()
        state = _normalize_state(ident.get("state"))
        issuer_raw = ident.get("issuer")
        issuer_norm = _normalize_issuer(issuer_raw)

        # --- 1. STATE LICENSE detection (highest priority). ---
        if _issuer_is_state_license(issuer_norm, desc_upper):
            key = _license_dedup_key(state, raw_id_str)
            # Pass 2: dispute against v02 licenses (already in
            # seen_license_keys from the v02 dedup above)?
            if key in seen_license_keys:
                # Determine whether the key is owned by a v02 entry
                # (dispute) or by an earlier OI entry (intra-OI dupe).
                owned_by_v02 = False
                kept_form = None
                for existing in deduped_v02_licenses:
                    if not isinstance(existing, dict):
                        continue
                    ex_state = existing.get("state")
                    ex_num = existing.get("number")
                    if ex_state and ex_num and _license_dedup_key(ex_state, ex_num) == key:
                        owned_by_v02 = True
                        kept_form = {"state": ex_state, "number": ex_num}
                        break
                if owned_by_v02:
                    counters.license_disputes += 1
                    dropped_form = {"state": state, "number": raw_id_str}
                    log.warning(
                        "license dispute: npi=%s kept=%r dropped=%r normalized_key=%r",
                        npi, kept_form, dropped_form, key,
                    )
                else:
                    # Pass 3: intra-OI dupe. Silent.
                    counters.oi_license_dupes_dropped += 1
                continue
            seen_license_keys.add(key)
            final_licenses.append({"state": state, "number": raw_id_str})
            counters.state_license_added += 1
            continue

        # --- 2. INSURANCE detection. ---
        if type_code in ("02", "05") or "MEDICAID" in desc_upper:
            brand = (f"State Medicaid Agency — {state}" if state
                     else "State Medicaid Agency")
            entry = {
                "insurance_type": "Medicaid",
                "brand": brand,
                "state": state,
                "raw_value": raw_id_str,
            }
            key = ("MEDICAID", brand.upper().strip(), state, raw_id_str)
            if key in seen_insurance_keys:
                continue
            seen_insurance_keys.add(key)
            insurance.append(entry)
            counters.medicaid_insurance_added += 1
            continue

        if type_code == "04" or "UPIN" in desc_upper:
            entry = {
                "insurance_type": "Medicare",
                "brand": "CMS",
                "state": state,
                "raw_value": raw_id_str,
            }
            key = ("MEDICARE", "CMS", state, raw_id_str)
            if key in seen_insurance_keys:
                continue
            seen_insurance_keys.add(key)
            insurance.append(entry)
            counters.upin_insurance_added += 1
            continue

        if _issuer_is_commercial_carrier(issuer_norm):
            brand = str(issuer_raw).strip()
            entry = {
                "insurance_type": "Commercial",
                "brand": brand,
                "state": state,
                "raw_value": raw_id_str,
            }
            key = ("COMMERCIAL", brand.upper().strip(), state, raw_id_str)
            if key in seen_insurance_keys:
                continue
            seen_insurance_keys.add(key)
            insurance.append(entry)
            counters.commercial_insurance_added += 1
            continue

        # --- 3. Fall-through. ---
        if issuer_raw is not None and str(issuer_raw).strip() != "":
            brand = str(issuer_raw).strip()
            entry = {
                "insurance_type": "",
                "brand": brand,
                "state": state,
                "raw_value": raw_id_str,
            }
            key = ("", brand.upper().strip(), state, raw_id_str)
            if key in seen_insurance_keys:
                continue
            seen_insurance_keys.add(key)
            insurance.append(entry)
            counters.unclassified_insurance_added += 1
            continue

        # No issuer and no matching rule: count by type_code and drop.
        counters.dropped_by_type_code[type_code or "<missing>"] += 1

    # Final-dedup assertion: NO duplicates may remain in licenses[].
    final_keys: set[tuple[str, str]] = set()
    for lic in final_licenses:
        if not isinstance(lic, dict):
            continue
        st = lic.get("state")
        num = lic.get("number")
        if st and num:
            key = _license_dedup_key(st, num)
            if key in final_keys:
                raise AssertionError(
                    f"duplicate license key after dedup: npi={npi} key={key!r}"
                )
            final_keys.add(key)

    out = dict(doc)
    if final_licenses:
        out["licenses"] = final_licenses
    else:
        out.pop("licenses", None)

    if insurance:
        out["insurance"] = insurance
    else:
        out.pop("insurance", None)

    out.pop("carriers", None)
    out.pop("_id", None)
    return out


# ---------------------------------------------------------------------------
# Run loop
# ---------------------------------------------------------------------------

def _flush(dst, batch: list[ReplaceOne], counters: TransformCounters) -> None:
    if not batch:
        return
    try:
        result = dst.bulk_write(batch, ordered=False)
    except BulkWriteError as exc:
        details = exc.details or {}
        counters.failed_docs += len(details.get("writeErrors", []))
        n_ok = (details.get("nUpserted", 0) + details.get("nModified", 0)
                + details.get("nInserted", 0))
        counters.written += n_ok
        log.error("bulk_write partial failure: errors=%d ok=%d",
                  len(details.get("writeErrors", [])), n_ok)
    else:
        counters.written += (
            (result.upserted_count or 0)
            + (result.modified_count or 0)
            + (result.inserted_count or 0)
        )


def run(limit: int | None, drop: bool, batch_size: int) -> int:
    conn = os.environ.get("MONGO_FRONTEND_connectionString")
    if not conn:
        log.error("MONGO_FRONTEND_connectionString is not set in environment")
        return 2

    _load_issuer_sets()

    client = MongoClient(conn, serverSelectionTimeoutMS=120_000)
    try:
        db = client[DB_NAME]
        src = db[SRC_COLLECTION]
        dst = db[DST_COLLECTION]

        if drop:
            log.info("--drop set; dropping %s.%s before processing",
                     DB_NAME, DST_COLLECTION)
            dst.drop()

        log.info("ensuring unique index on %s.npi", DST_COLLECTION)
        dst.create_index("npi", unique=True, name="npi_unique")

        counters = TransformCounters()

        log.info("scanning %s.%s -> %s.%s (limit=%s)",
                 DB_NAME, SRC_COLLECTION, DB_NAME, DST_COLLECTION,
                 limit if limit is not None else "ALL")

        cursor = src.find({}, no_cursor_timeout=True).batch_size(batch_size)
        if limit is not None and limit > 0:
            cursor = cursor.limit(limit)

        pending: list[ReplaceOne] = []
        try:
            for doc in cursor:
                counters.scanned += 1
                try:
                    out = transform_doc(doc, counters)
                except Exception as exc:
                    counters.failed_docs += 1
                    log.exception("transform failed for npi=%s: %s",
                                  doc.get("npi"), exc)
                    continue

                if out is None:
                    continue

                npi = out.get("npi")
                if not npi:
                    counters.failed_docs += 1
                    log.warning("doc missing npi, skipping: _id_was_present=%s",
                                "_id" in doc)
                    continue

                pending.append(ReplaceOne({"npi": npi}, out, upsert=True))
                if len(pending) >= batch_size:
                    _flush(dst, pending, counters)
                    pending = []

                if counters.scanned % PROGRESS_EVERY == 0:
                    log.info("progress: scanned=%d written=%d failed=%d",
                             counters.scanned, counters.written,
                             counters.failed_docs)
        finally:
            if pending:
                _flush(dst, pending, counters)
            cursor.close()

        log.info("=" * 70)
        log.info("RUN SUMMARY")
        log.info("  v02 docs scanned             : %d", counters.scanned)
        log.info("  v03 docs written             : %d", counters.written)
        log.info("  skipped out_of_scope         : %d", counters.skipped_out_of_scope)
        log.info("  v02 intra-license dupes      : %d", counters.v02_license_dupes_dropped)
        log.info("  OI intra-license dupes       : %d", counters.oi_license_dupes_dropped)
        log.info("  state-license added (OI)     : %d", counters.state_license_added)
        log.info("  Medicaid insurance added     : %d", counters.medicaid_insurance_added)
        log.info("  Medicare UPIN insurance      : %d", counters.upin_insurance_added)
        log.info("  Commercial insurance added   : %d", counters.commercial_insurance_added)
        log.info("  Unclassified insurance added : %d", counters.unclassified_insurance_added)
        log.info("  license disputes (OI dropped): %d", counters.license_disputes)
        log.info("  failed docs                  : %d", counters.failed_docs)
        if counters.dropped_by_reason:
            log.info("  dropped by out_of_scope reason:")
            for r, n in sorted(counters.dropped_by_reason.items()):
                log.info("     %-20s : %d", r, n)
        if counters.dropped_by_type_code:
            log.info("  dropped by type_code:")
            for tc, n in sorted(counters.dropped_by_type_code.items()):
                log.info("     %-12s : %d", tc, n)
        else:
            log.info("  dropped by type_code         : (none)")
        log.info("=" * 70)
        return 0
    finally:
        client.close()


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=("Build PublicHealthData.provider_v03 by mining "
                     "other_identifiers into licenses[] and insurance[] "
                     "on each provider_v02 document.")
    )
    ap.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N source documents (smoke run).",
    )
    ap.add_argument(
        "--drop",
        action="store_true",
        help="Drop provider_v03 before processing.",
    )
    ap.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=(f"bulk_write batch size (default {DEFAULT_BATCH_SIZE})."),
    )
    return ap.parse_args()


def main() -> int:
    _setup_logging()
    args = parse_args()
    return run(limit=args.limit, drop=args.drop, batch_size=args.batch_size)


if __name__ == "__main__":
    sys.exit(main())


# ──────────────────────────────────────────────────────────────────────
# CODE REVIEW (self-review by the implementer)
# ──────────────────────────────────────────────────────────────────────
# Finding 1: HIGH — STATE LICENSE detection now precedes INSURANCE
#   detection. The prior version mis-routed every non-empty issuer to
#   Commercial insurance, which sent ~4,451 distinct state-board /
#   state-DOH issuer strings (and the ~14,704 "LICENSE" literal-token
#   entries) into insurance[] instead of licenses[]. FIXED by seeding
#   the classifier from the workbook's STATE_BOARD set + the canonical
#   board phrases + the literal-license tokens + the
#   type_description "LICENSE" word check.
#
# Finding 2: HIGH — OI-derived licenses are minted with EXACTLY two
#   fields, {state, number}, per the operator directive. No source
#   annotation, no classification metadata, no issuer. This matches
#   the v02 schema's license shape. FIXED.
#
# Finding 3: HIGH — Three-pass dedup with a post-emit assertion
#   guarantees no two licenses[] entries share the (state,
#   normalized_number) key: pass 1 dedups v02 internal dupes silently;
#   pass 2 logs a "license dispute" when an OI candidate collides with
#   a v02 entry and drops the OI candidate; pass 3 silently drops
#   intra-OI dupes; the final assertion raises if any duplicate slips
#   through (transform bug). FIXED.
#
# Finding 4: HIGH — Out-of-scope documents are dropped entirely (no
#   write to v03), with a per-reason breakdown in the run summary.
#   Preserved verbatim from the prior version.
#
# Finding 5: HIGH — Unclassified insurance fall-through: OI entries
#   with a non-empty issuer that match no STATE_BOARD, no Medicaid /
#   Medicare / commercial-carrier rule emit to insurance[] with
#   insurance_type="". Honest "we couldn't classify" signal; the
#   operator can audit these rows later. FIXED.
#
# Finding 6: MEDIUM — Issuer normalization matches the workbook
#   exactly: uppercase, collapse internal whitespace, strip leading
#   /trailing whitespace and trailing/leading punctuation
#   (`,.;:- `). Verified on workbook examples ("BCBS," → "BCBS",
#   "BLUE  CROSS BLUE SHIELD" → "BLUE CROSS BLUE SHIELD",
#   "RAILROAD MEDICARE  -" → "RAILROAD MEDICARE"). No regex used
#   (Rule-008 statement 4 compliant).
#
# Finding 7: MEDIUM — _load_issuer_sets() reads the workbook at
#   startup (both from run() and from the smoke wrapper) and caches
#   the STATE_BOARD and COMMERCIAL_CARRIER sets module-level. If the
#   workbook is missing, the classifier falls back to phrase-only
#   matching with a logged warning. The phrase rules cover the
#   high-frequency cases (BOARD OF *, DEPARTMENT OF HEALTH,
#   AETNA/BCBS/CIGNA/...) so coverage degrades gracefully rather
#   than crashing.
#
# Finding 8: MEDIUM — Empty-array rule preserved: licenses and
#   insurance are OMITTED entirely (not emitted as []) when empty.
#   Downstream consumers that key on doc.get("licenses", []) keep
#   working; consumers that key on field presence change behavior.
#   Spec-mandated.
#
# Finding 9: MEDIUM — License-dispute warning identifies the "kept"
#   entry by walking deduped_v02_licenses linearly. Bounded by the
#   per-provider license count (tens at most); not a hot path. KEPT
#   as the simplest correct rendering.
#
# Finding 10: LOW — `carriers[]` is unconditionally stripped from
#   the v03 output via `out.pop("carriers", None)`. Today's v02 does
#   not carry one; the pop is defensive. KEPT.
#
# Finding 11: LOW — Insurance dedup compares brand case-insensitively
#   (key uses brand.upper().strip()) but the stored brand preserves
#   case. First-occurrence case wins on collision. KEPT.
#
# Finding 12: LOW — `_id` is stripped before write; Mongo assigns a
#   fresh ObjectId on insert. Stable downstream key is `npi`, which
#   is uniquely indexed. KEPT.
#
# Finding 13: LOW — All v02 fields other than `_id`, `licenses`,
#   `insurance`, and `carriers` pass through verbatim via
#   `dict(doc)`. No normalization, stripping, or invention of any
#   other field. KEPT.
#
# Severity totals: HIGH = 5 (all fixed) · MEDIUM = 4 (all per-spec)
#                  LOW   = 4 (all kept with rationale)
# ──────────────────────────────────────────────────────────────────────
