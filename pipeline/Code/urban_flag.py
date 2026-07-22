# Copyright (c) 2026 ChatHealthy.ai LLC. All rights reserved.
# Licensed under the FindCare Evaluation License (FEL-1.0).

"""USDA RUCC source-gather job.

Realizes EPIC-010-F-103-S-005 source step. Runs in the source-gather
phase. Discovers the latest USDA RUCC xlsx URL via the AI agent,
caches the xlsx in the blob container, parses it once into a dict
of {county_fips: urban_bool}, and writes that dict as `rucc.json` in
the same blob container — the durable shard every chunk worker reads.

RUCC 1-3 -> urban=True, RUCC 4-9 -> urban=False.

Per-record stamping (addresses[].county.urban) is owned by
process_assignment_activity._stamp_urban, which loads rucc.json once
per worker process via _load_rucc.
"""

from __future__ import annotations
from chathealthy_frontend_lib.logging_service import ChatHealthyLoggingService

import io
import json

import os

import requests


_log = ChatHealthyLoggingService()

_USDA_RUCC_INDEX_URL = "https://www.ers.usda.gov/data-products/rural-urban-continuum-codes/"

RUCC_JSON_BLOB_NAME = "rucc.json"
URBAN_RUCC_MAX = 3  # RUCC 1..3 are urban; 4..9 are rural


def cache_rucc_to_blob(blob_container: str) -> dict:
    """One-shot source-gather work: agent-discover URL, cache xlsx in
    blob, parse, write rucc.json. Returns metric.

    The caller (gather_rucc_activity_fn) needs no further state. Every
    chunk worker reads rucc.json from the blob on first use.
    """
    from blob_client import get_blob_service
    from openpyxl import load_workbook
    from source_url_discovery import find_latest_data_url

    rucc_url = find_latest_data_url(
        source_name="usda_rucc",
        page_url=_USDA_RUCC_INDEX_URL,
        instructions=(
            "Find the URL of the latest USDA Rural-Urban Continuum "
            "Codes (RUCC) spreadsheet file. Rules: (a) Look for the "
            "most recent year's RUCC file (filename typically "
            "`Ruralurbancontinuumcodes<YEAR>.xlsx` or similar). "
            "(b) MUST be an XLSX or CSV, not PDF. (c) Pick the file "
            "that contains the actual county-level RUCC codes 1-9, "
            "not documentation. (d) Return only the absolute URL."
        ),
    )

    container = get_blob_service().get_container_client(blob_container)

    xlsx_blob_name = "usda_" + rucc_url.rsplit("/", 1)[-1].split("?")[0]
    xlsx_blob = container.get_blob_client(xlsx_blob_name)
    try:
        data = xlsx_blob.download_blob().readall()
        _log.info("urban_flag: blob hit %s (%d bytes)", xlsx_blob_name, len(data))
    except Exception as exc:
        _log.info("urban_flag: blob miss (%s) — fetching USDA", exc)
        resp = requests.get(rucc_url, timeout=120)
        resp.raise_for_status()
        data = resp.content
        try:
            xlsx_blob.upload_blob(data, overwrite=True)
        except Exception as upload_exc:
            _log.warning("urban_flag: xlsx blob cache failed (%s)", upload_exc)

    fips_to_urban: dict[str, bool] = {}
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            rows = wb[sheet_name].iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            norm_header = [str(h or "").strip().lower() for h in header]
            fips_idx = None
            for cand in ("fips", "fips_code", "fipscode", "fips_txt"):
                if cand in norm_header:
                    fips_idx = norm_header.index(cand)
                    break
            rucc_idx = None
            for i, h in enumerate(norm_header):
                if h.startswith("rucc_") or h == "rucc" or "ruralurbancontinuum" in h:
                    rucc_idx = i
                    break
            if fips_idx is None or rucc_idx is None:
                continue
            _log.info(
                "urban_flag: parsing sheet %r (fips col %d, rucc col %d)",
                sheet_name, fips_idx, rucc_idx,
            )
            for row in rows:
                try:
                    fips = str(row[fips_idx]).strip().zfill(5)
                    rucc = int(row[rucc_idx])
                except (TypeError, ValueError, IndexError):
                    continue
                if len(fips) != 5 or rucc < 1 or rucc > 9:
                    continue
                fips_to_urban[fips] = rucc <= URBAN_RUCC_MAX
            break  # first matching sheet wins
    finally:
        wb.close()

    json_blob = container.get_blob_client(RUCC_JSON_BLOB_NAME)
    json_blob.upload_blob(
        json.dumps(fips_to_urban, separators=(",", ":")).encode("utf-8"),
        overwrite=True,
    )
    _log.info(
        "urban_flag: wrote %s (%d county entries) to %s",
        RUCC_JSON_BLOB_NAME, len(fips_to_urban), blob_container,
    )
    return {
        "rucc_source_url": rucc_url,
        "xlsx_blob": xlsx_blob_name,
        "json_blob": RUCC_JSON_BLOB_NAME,
        "rucc_loaded_counties": len(fips_to_urban),
    }
