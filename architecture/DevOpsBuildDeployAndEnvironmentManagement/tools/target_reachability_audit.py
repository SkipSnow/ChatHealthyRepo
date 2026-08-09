#!/usr/bin/env python3
"""Semantic reachability audit for deployment targets.

Computes minimum file sets per target from entry points and Dockerfile COPY
roots, compares against deployment_architecture.json files[] and
huggingface_source_set wholesale trees, and emits a refactor spreadsheet.

Usage (from repo root):
  python architecture/DevOpsBuildDeployAndEnvironmentManagement/target_reachability_audit.py

Outputs (under _oneshots/):
  target_deployment_refactor_<date>.xlsx + .csv
  target_deployment_proposed_files_<date>.json  (files[] drafts for 5 new targets)
  BUILD_DEPLOY_MINIMUM_FACTS_PROPOSAL.md (copy in DevOps folder)
"""
from __future__ import annotations

import ast
import csv
import hashlib
import json
import re
import subprocess
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parents[3]
DEVOPS = Path(__file__).resolve().parent.parent
DEPLOY_JSON = ROOT / "brain" / "machine_artifacts" / "content" / "deployment_architecture.json"
ENG_RULES = ROOT / "brain" / "machine_artifacts" / "content" / "engineering_rules.json"
OUT_DIR = ROOT / "_oneshots"
TODAY = date.today().isoformat()

FEATURE_DEFAULT = "EPIC-008-F-012"

# ── Target definitions (entrances + structural roots) ─────────────────────

@dataclass
class TargetSpec:
    target_id: str
    cloud_mirror: str | None
    entry_modules: list[str]
    dockerfile: str | None
    docker_copy_roots: list[str]
    static_roots: list[str] = field(default_factory=list)
    node_address: str = "localhost"
    deployable: bool = True


NEW_LOCAL_TARGETS: list[TargetSpec] = [
    TargetSpec(
        target_id="target_docker_local_findcare",
        cloud_mirror="target_hf_space_findcare_backend",
        entry_modules=["Code/ConversationalUX/FindCareChat/backend/app.py"],
        dockerfile="DevOps/FindCareBackend/Dockerfile",
        docker_copy_roots=[
            "DevOps/FindCareBackend/requirements.txt",
            "Code/ConversationalUX/FindCareChat/backend",
            "Code/Shared",
            "Code/ConversationalUX/ChatHealthyWhoAmIChat/me",
            "brain",
            "FindCare",
            "FrontEndApplicationLib",
        ],
        node_address="localhost:7860",
    ),
    TargetSpec(
        target_id="target_docker_local_evaluatecare",
        cloud_mirror="target_hf_space_evaluatecare_backend",
        entry_modules=["evaluateCare/Code/app.py"],
        dockerfile="evaluateCare/Code/Dockerfile",
        docker_copy_roots=[
            "evaluateCare/Code/requirements.txt",
            "evaluateCare/Code/app.py",
            "evaluateCare/Code/healthcheck",
            "evaluateCare/Code/displayChrome",
            "evaluateCare/Code/security",
            "evaluateCare/Code/externalInterface",
            "FrontEndApplicationLib",
        ],
        node_address="localhost:8001",
    ),
    TargetSpec(
        target_id="target_docker_local_sharedservices",
        cloud_mirror="target_hf_space_shared_services",
        entry_modules=["sharedServices/Code/app.py"],
        dockerfile="sharedServices/Code/Dockerfile",
        docker_copy_roots=[
            "sharedServices/Code/requirements.txt",
            "sharedServices/Code/app.py",
            "sharedServices/Code/healthcheck",
            "sharedServices/Code/displayChrome",
            "sharedServices/Code/authentication",
            "sharedServices/Code/SpecialtyFilter",
            "sharedServices/Code/ClinicalTrials",
            "sharedServices/Code/AboutChatHealthy",
            "sharedServices/Code/externalInterface",
            "sharedServices/Code/secretsManager",
            "sharedServices/Code/UtteranceManager",
            "sharedServices/Code/NonsenseTool",
            "sharedServices/Code/CloseConnection200Tool",
            "sharedServices/Code/LockoutTool",
            "sharedServices/Code/ProviderSelection",
            "FrontEndApplicationLib",
        ],
        node_address="localhost:8002",
    ),
    TargetSpec(
        target_id="target_docker_local_website",
        cloud_mirror="target_cloudflare_pages_website",
        entry_modules=[
            "architecture/DevOpsBuildDeployAndEnvironmentManagement/_start_website.py",
        ],
        dockerfile="architecture/DevOpsBuildDeployAndEnvironmentManagement/WebsiteWrapper/Dockerfile",
        docker_copy_roots=[
            "architecture/DevOpsBuildDeployAndEnvironmentManagement/_start_website.py",
            "architecture/DevOpsBuildDeployAndEnvironmentManagement/WebsiteWrapper/Dockerfile",
        ],
        static_roots=["Website"],
        node_address="localhost:443",
    ),
    TargetSpec(
        target_id="target_host_local_workstation",
        cloud_mirror=None,
        entry_modules=[
            "Code/Shared/ops/tools/chathealthy_devops_boot.py",
            "architecture/DevOpsBuildDeployAndEnvironmentManagement/build_chathealthy.py",
            "architecture/DevOpsBuildDeployAndEnvironmentManagement/deploy_chathealthy.py",
        ],
        dockerfile=None,
        docker_copy_roots=[],
        deployable=False,
        node_address="localhost/workstation",
    ),
]

WORKSTATION_PREFIXES = (
    "architecture/DevOpsBuildDeployAndEnvironmentManagement/",
    "architecture/ClaudeCodeConversationPersistence/",
    "brain/machine_artifacts/content/",
    "Code/Shared/ops/tools/",
    "Code/Shared/ops/certs/",
    ".claude/",
    "Website/schemas/",
)

JUNK_SUFFIXES = (".pyc", ".pyo", ".DS_Store")
JUNK_PATH_PARTS = ("__pycache__", ".pytest_cache", "node_modules", ".venv", "localBuild/")


def git_tracked_files() -> list[str]:
    out = subprocess.run(
        ["git", "ls-files"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    return sorted(line.strip().replace("\\", "/") for line in out.stdout.splitlines() if line.strip())


def load_manifest() -> dict:
    return json.loads(DEPLOY_JSON.read_text(encoding="utf-8"))


def manifest_index(manifest: dict) -> dict[str, dict]:
    return {r["target_id"]: r for r in manifest["DeploymentTargetRecord"]}


def files_in_manifest_record(rec: dict) -> set[str]:
    return {f["source_location"] for f in rec.get("files", [])}


def wholesale_from_record(rec: dict) -> set[str]:
    out: set[str] = set()
    for item in rec.get("huggingface_source_set") or []:
        src = item["src"].replace("\\", "/").rstrip("/")
        p = ROOT / src
        if p.is_file():
            out.add(src)
        elif p.is_dir():
            for f in p.rglob("*"):
                if f.is_file():
                    out.add(f.relative_to(ROOT).as_posix())
    return out


def expand_roots(roots: Iterable[str]) -> set[str]:
    out: set[str] = set()
    for root in roots:
        root = root.replace("\\", "/")
        p = ROOT / root
        if p.is_file():
            out.add(root)
        elif p.is_dir():
            for f in p.rglob("*"):
                if f.is_file():
                    out.add(f.relative_to(ROOT).as_posix())
    return out


def content_hash(rel: str) -> str:
    p = ROOT / rel
    if not p.is_file():
        return ""
    raw = p.read_bytes().replace(b"\r\n", b"\n")
    return hashlib.sha256(raw).hexdigest()


def handler_type_for(rel: str) -> str:
    low = rel.lower()
    if low.endswith(".py"):
        if "/tests/" in low or low.endswith("_test.py") or low.startswith("test_"):
            return "python_test"
        return "python_source"
    if low.endswith((".html", ".htm", ".css", ".js", ".tsx", ".jsx", ".svg", ".png", ".jpg", ".ico", ".woff", ".woff2")):
        return "static_asset"
    if low.endswith(".json"):
        return "json"
    if low.endswith((".md", ".txt")):
        return "markdown" if low.endswith(".md") else "static_asset"
    if low.endswith(".yml") or low.endswith(".yaml"):
        return "workflow_yaml"
    if "dockerfile" in low.split("/")[-1].lower():
        return "dockerfile"
    if low.endswith(".crt") or low.endswith(".key") or low.endswith(".pem"):
        return "cert_or_key"
    if low.endswith("requirements.txt"):
        return "python_requirements"
    return "static_asset"


def is_junk(rel: str) -> bool:
    if any(part in rel for part in JUNK_PATH_PARTS):
        return True
    return rel.endswith(JUNK_SUFFIXES)


def hook_and_rule_paths() -> set[str]:
    out: set[str] = set()
    if ENG_RULES.is_file():
        data = json.loads(ENG_RULES.read_text(encoding="utf-8"))
        for rule in data.get("rules", {}).get("rule", []):
            for enf in rule.get("enforcements", {}).get("enforcement", []):
                p = enf.get("executable_path", "").replace("\\", "/")
                if p:
                    out.add(p)
    settings = ROOT / ".claude" / "settings.json"
    if settings.is_file():
        out.add(".claude/settings.json")
    return out


def path_roots_for_spec(spec: TargetSpec) -> list[Path]:
    roots: list[Path] = []
    for r in spec.docker_copy_roots:
        p = ROOT / r
        if p.is_dir():
            roots.append(p)
        elif p.parent != ROOT:
            roots.append(p.parent)
    # FindCare runtime sys.path roots
    if spec.target_id == "target_docker_local_findcare":
        roots.extend([
            ROOT / "FindCare",
            ROOT / "Code" / "Shared",
            ROOT / "Code" / "ConversationalUX" / "FindCareChat" / "backend",
            ROOT / "FrontEndApplicationLib",
            ROOT / "brain" / "machine_artifacts" / "content",
        ])
    elif spec.target_id == "target_docker_local_evaluatecare":
        roots.append(ROOT / "evaluateCare" / "Code")
        roots.append(ROOT / "FrontEndApplicationLib")
    elif spec.target_id == "target_docker_local_sharedservices":
        roots.append(ROOT / "sharedServices" / "Code")
        roots.append(ROOT / "FrontEndApplicationLib")
    elif spec.target_id == "target_host_local_workstation":
        roots.append(DEVOPS)
        roots.append(ROOT / "Code" / "Shared" / "ops" / "tools")
    dedup: list[Path] = []
    seen: set[str] = set()
    for r in roots:
        key = str(r.resolve())
        if key not in seen:
            seen.add(key)
            dedup.append(r)
    return dedup


def module_to_relpaths(module: str, roots: list[Path]) -> list[str]:
    """Resolve import module name to repo-relative paths."""
    found: list[str] = []
    search_roots: list[Path] = []
    for root in roots:
        if root.exists():
            search_roots.append(root)
            src = root / "src"
            if src.is_dir():
                search_roots.append(src)

    parts = module.split(".")
    for root in search_roots:
        # package/module path: a.b.c -> root/a/b/c.py or root/a/b/c/__init__.py
        mod_path = root.joinpath(*parts)
        py_file = mod_path.with_suffix(".py")
        if py_file.is_file():
            found.append(py_file.relative_to(ROOT).as_posix())
            continue
        init_file = mod_path / "__init__.py"
        if init_file.is_file():
            found.append(init_file.relative_to(ROOT).as_posix())
            continue
        # single-segment fallback at root
        if len(parts) == 1:
            one = root / f"{parts[0]}.py"
            if one.is_file():
                found.append(one.relative_to(ROOT).as_posix())
    return list(dict.fromkeys(found))


def resolve_import_from(node: ast.ImportFrom, current: Path, roots: list[Path]) -> list[str]:
    search_dirs: list[Path] = []
    d = current.parent
    while True:
        search_dirs.append(d)
        if d == ROOT or d.parent == d:
            break
        d = d.parent

    if node.level and node.level > 0:
        base = current.parent
        for _ in range(node.level - 1):
            base = base.parent
        pkg_parts = (node.module or "").split(".") if node.module else []
        out: list[str] = []
        for alias in node.names:
            if alias.name == "*":
                continue
            rel_parts = list(base.relative_to(ROOT).parts) + pkg_parts + [f"{alias.name}.py"]
            cand = ROOT.joinpath(*rel_parts)
            if cand.is_file():
                out.append(cand.relative_to(ROOT).as_posix())
            else:
                init = cand.parent / alias.name / "__init__.py"
                if init.is_file():
                    out.append(init.relative_to(ROOT).as_posix())
        return out

    mod = node.module or ""
    out: list[str] = []
    for alias in node.names:
        if alias.name == "*":
            if mod:
                out.extend(module_to_relpaths(mod, search_dirs + roots))
            continue
        full = f"{mod}.{alias.name}" if mod else alias.name
        hits = module_to_relpaths(full, search_dirs + roots)
        if not hits and mod:
            hits = module_to_relpaths(mod, search_dirs + roots)
        out.extend(hits)
    return list(dict.fromkeys(out))


def runtime_json_refs(spec: TargetSpec) -> set[str]:
    """Known runtime reads of brain JSON and build artifacts."""
    out: set[str] = set()
    if spec.target_id in (
        "target_docker_local_findcare",
        "target_hf_space_findcare_backend",
    ):
        for name in (
            "agile_backlog.json",
            "deployment_architecture.json",
            "engineering_rules.json",
            "bugs.json",
            "risk_acceptance.json",
        ):
            p = ROOT / "brain" / "machine_artifacts" / "content" / name
            if p.is_file():
                out.add(p.relative_to(ROOT).as_posix())
        # PromptSystemMaker and runtime_data_collections pull brain JSON
        psm = ROOT / "Code" / "Shared" / "prompt_system_maker.py"
        if psm.is_file():
            out.add(psm.relative_to(ROOT).as_posix())
        rdc = ROOT / "FrontEndApplicationLib" / "src" / "chathealthy_frontend_lib" / "runtime_data_collections.py"
        if rdc.is_file():
            out.add(rdc.relative_to(ROOT).as_posix())
    return out


def literal_path_refs(path: Path, seen: set[str]) -> list[str]:
    """Extract repo-relative paths from string literals in Python source."""
    out: list[str] = []
    try:
        tree = ast.parse(path.read_text(encoding="utf-8", errors="replace"))
    except SyntaxError:
        return out
    for node in ast.walk(tree):
        if isinstance(node, ast.Constant) and isinstance(node.value, str):
            val = node.value.replace("\\", "/")
            if val.startswith(("brain/", "Code/", "FindCare/", "Website/")):
                p = ROOT / val
                if p.is_file():
                    rel = p.relative_to(ROOT).as_posix()
                    if rel not in seen:
                        out.append(rel)
    return out


def semantic_reachable(spec: TargetSpec) -> set[str]:
    roots = path_roots_for_spec(spec)
    queue: list[str] = list(spec.entry_modules)
    seen: set[str] = set()
    while queue:
        rel = queue.pop()
        rel = rel.replace("\\", "/")
        if rel in seen or is_junk(rel):
            continue
        seen.add(rel)
        path = ROOT / rel
        if not path.is_file():
            continue
        if path.suffix == ".py":
            try:
                tree = ast.parse(path.read_text(encoding="utf-8", errors="replace"))
            except SyntaxError:
                continue
            for node in ast.walk(tree):
                if isinstance(node, ast.Import):
                    for alias in node.names:
                        for hit in module_to_relpaths(alias.name, roots):
                            if hit not in seen:
                                queue.append(hit)
                        for hit in module_to_relpaths(alias.name, [path.parent]):
                            if hit not in seen:
                                queue.append(hit)
                elif isinstance(node, ast.ImportFrom):
                    for hit in resolve_import_from(node, path, roots):
                        if hit not in seen:
                            queue.append(hit)
            for hit in literal_path_refs(path, seen):
                queue.append(hit)
        elif path.suffix in (".html", ".js", ".tsx", ".jsx"):
            text = path.read_text(encoding="utf-8", errors="replace")
            for m in re.finditer(r'''(?:src|href)=["']([^"']+)["']''', text):
                ref = m.group(1).split("?")[0]
                if ref.startswith(("http://", "https://", "//", "#", "data:")):
                    continue
                resolved = (path.parent / ref).resolve()
                try:
                    rel_ref = resolved.relative_to(ROOT.resolve()).as_posix()
                except ValueError:
                    continue
                if rel_ref not in seen:
                    queue.append(rel_ref)
    seen |= runtime_json_refs(spec)
    return seen


def static_reachable(spec: TargetSpec) -> set[str]:
    out: set[str] = set()
    for root in spec.static_roots:
        out |= expand_roots([root])
    # Website build dir is substituted copy — treat Website/ as source SSOT
    return out


def workstation_files(all_tracked: set[str], deploy_union: set[str]) -> set[str]:
    out: set[str] = set()
    out |= hook_and_rule_paths()
    for rel in all_tracked:
        if rel in deploy_union:
            continue
        if any(rel.startswith(p) for p in WORKSTATION_PREFIXES):
            out.add(rel)
        elif rel.startswith("architecture/") and rel.endswith(".py"):
            out.add(rel)
    # Entire DevOps substrate already in target_host_local_deploy_tool manifest
    idx = manifest_index(load_manifest())
    if "target_host_local_deploy_tool" in idx:
        out |= files_in_manifest_record(idx["target_host_local_deploy_tool"])
    return out


def file_composition_entries(paths: Iterable[str]) -> list[dict]:
    rows = []
    for rel in sorted(set(paths)):
        if is_junk(rel) or not (ROOT / rel).is_file():
            continue
        rows.append({
            "feature_id": FEATURE_DEFAULT,
            "source_location": rel,
            "handler_type": handler_type_for(rel),
            "disposition": "referenced",
            "content_hash": content_hash(rel),
        })
    return rows


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()
    idx = manifest_index(manifest)
    tracked = git_tracked_files()
    tracked_set = set(tracked)

    specs = list(NEW_LOCAL_TARGETS)
    cloud_ids = [
        "target_cloudflare_pages_website",
        "target_hf_space_findcare_backend",
        "target_hf_space_evaluatecare_backend",
        "target_hf_space_shared_services",
        "target_host_local_deploy_tool",
    ]

    structural: dict[str, set[str]] = {}
    semantic: dict[str, set[str]] = {}
    declared: dict[str, set[str]] = {}
    wholesale: dict[str, set[str]] = {}

    for spec in specs:
        structural[spec.target_id] = expand_roots(spec.docker_copy_roots) | static_reachable(spec)
        semantic[spec.target_id] = semantic_reachable(spec) | static_reachable(spec)
        if spec.target_id == "target_host_local_workstation":
            deploy_union = set()
            for s in specs:
                if s.deployable:
                    deploy_union |= semantic.get(s.target_id, set())
            for cid in cloud_ids:
                if cid in idx:
                    deploy_union |= files_in_manifest_record(idx[cid])
            semantic[spec.target_id] = workstation_files(tracked_set, deploy_union)
            structural[spec.target_id] = set(semantic[spec.target_id])

    for tid in cloud_ids:
        if tid in idx:
            declared[tid] = files_in_manifest_record(idx[tid])
            wholesale[tid] = wholesale_from_record(idx[tid])

    for spec in specs:
        if spec.cloud_mirror and spec.cloud_mirror in idx:
            declared[spec.target_id] = declared.get(spec.cloud_mirror, set())
            wholesale[spec.target_id] = wholesale.get(spec.cloud_mirror, set())

    all_target_ids = sorted(
        set(declared) | set(structural) | set(semantic) | {s.target_id for s in specs}
    )

    # Spreadsheet rows
    rows_out: list[dict] = []
    for rel in tracked:
        if is_junk(rel):
            continue
        sem_hits = [t for t in specs if rel in semantic.get(t.target_id, set())]
        struct_hits = [t for t in specs if rel in structural.get(t.target_id, set())]
        decl_hits = [t for t in all_target_ids if rel in declared.get(t, set())]
        whole_hits = [t for t in all_target_ids if rel in wholesale.get(t, set())]

        # Proposed: semantic minimum for new local targets; declared for cloud
        proposed: list[str] = []
        for spec in specs:
            if rel in semantic.get(spec.target_id, set()):
                proposed.append(spec.target_id)
        for tid in cloud_ids:
            if tid in idx and rel in semantic.get(tid, set()):
                proposed.append(tid)

        bloat_targets = [
            t for t in whole_hits
            if rel in wholesale.get(t, set()) and rel not in declared.get(t, set())
        ]
        gap = [t for t in proposed if t not in decl_hits]

        if rel in semantic.get("target_host_local_workstation", set()):
            disposition = "WORKSTATION"
        elif bloat_targets:
            disposition = "WHOLESALE_BLOAT"
        elif sem_hits and not decl_hits:
            disposition = "SEMANTIC_ONLY"
        elif decl_hits and not sem_hits:
            disposition = "DECLARED_ONLY"
        elif decl_hits or sem_hits:
            disposition = "LIVE"
        elif whole_hits:
            disposition = "WHOLESALE_ONLY"
        elif any(rel.startswith(p) for p in WORKSTATION_PREFIXES):
            disposition = "WORKSTATION_CANDIDATE"
        else:
            disposition = "UNREACHABLE"

        rows_out.append({
            "file_path": rel,
            "disposition": disposition,
            "semantic_targets": ";".join(sorted({s.target_id for s in specs if rel in semantic.get(s.target_id, set())})),
            "structural_targets": ";".join(sorted({s.target_id for s in specs if rel in structural.get(s.target_id, set())})),
            "structural_only": ";".join(sorted(
                s.target_id for s in specs
                if rel in structural.get(s.target_id, set()) and rel not in semantic.get(s.target_id, set())
            )),
            "declared_in_files": ";".join(sorted(decl_hits)),
            "wholesale_copy": ";".join(sorted(whole_hits)),
            "proposed_targets": ";".join(sorted(set(proposed))),
            "gap_vs_declared": ";".join(sorted(gap)),
            "bloat_vs_files": ";".join(sorted(bloat_targets)),
            "handler_type": handler_type_for(rel),
            "content_hash": content_hash(rel),
        })

    csv_path = OUT_DIR / f"target_deployment_refactor_{TODAY}.csv"
    xlsx_path = OUT_DIR / f"target_deployment_refactor_{TODAY}.xlsx"
    fieldnames = list(rows_out[0].keys()) if rows_out else []

    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows_out)

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "files"
        ws.append(fieldnames)
        for row in rows_out:
            ws.append([row[k] for k in fieldnames])
        for i, _ in enumerate(fieldnames, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(40, max(12, len(fieldnames[i - 1]) + 2))
        # Summary sheet
        sm = wb.create_sheet("summary")
        sm.append(["target_id", "semantic_count", "structural_count", "declared_count", "wholesale_count", "bloat_struct_minus_semantic"])
        for spec in specs:
            tid = spec.target_id
            sm.append([
                tid,
                len(semantic.get(tid, set())),
                len(structural.get(tid, set())),
                len(declared.get(tid, set())),
                len(wholesale.get(tid, set())),
                len(structural.get(tid, set()) - semantic.get(tid, set())),
            ])
        for tid in cloud_ids:
            if tid not in idx:
                continue
            sm.append([
                tid,
                "",
                "",
                len(declared.get(tid, set())),
                len(wholesale.get(tid, set())),
                len(wholesale.get(tid, set()) - declared.get(tid, set())),
            ])
        wb.save(xlsx_path)
    except ImportError:
        xlsx_path = None

    # Proposed files[] for 5 new targets
    proposed_records: dict[str, list[dict]] = {}
    for spec in specs:
        proposed_records[spec.target_id] = file_composition_entries(semantic.get(spec.target_id, set()))

    proposed_path = OUT_DIR / f"target_deployment_proposed_files_{TODAY}.json"
    proposed_path.write_text(
        json.dumps(proposed_records, indent=2),
        encoding="utf-8",
    )

    stats = summarize(rows_out, specs, semantic, structural, declared, wholesale)
    proposal_md = render_proposal(stats, specs, proposed_path, csv_path, xlsx_path)
    proposal_devops = DEVOPS / "BUILD_DEPLOY_MINIMUM_FACTS_PROPOSAL.md"
    proposal_devops.write_text(proposal_md, encoding="utf-8")
    (OUT_DIR / "BUILD_DEPLOY_MINIMUM_FACTS_PROPOSAL.md").write_text(proposal_md, encoding="utf-8")

    print(proposal_md[:2000])
    print("\n---\n")
    print(f"CSV:  {csv_path}")
    if xlsx_path:
        print(f"XLSX: {xlsx_path}")
    print(f"Proposed files[] JSON: {proposed_path}")
    print(f"Proposal: {proposal_devops}")


def summarize(rows, specs, semantic, structural, declared, wholesale) -> dict:
    disp = defaultdict(int)
    for r in rows:
        disp[r["disposition"]] += 1
    return {
        "tracked_files": len(rows),
        "disposition": dict(disp),
        "targets": {
            s.target_id: {
                "semantic": len(semantic.get(s.target_id, set())),
                "structural": len(structural.get(s.target_id, set())),
                "declared_mirror": len(declared.get(s.target_id, set())),
                "wholesale_mirror": len(wholesale.get(s.target_id, set())),
            }
            for s in specs
        },
    }


def render_proposal(stats: dict, specs: list[TargetSpec], proposed_path: Path, csv_path: Path, xlsx_path: Path | None) -> str:
    lines = [
        f"# Build / Deploy Minimum Facts Proposal ({TODAY})",
        "",
        "## Goal",
        "",
        "Stop BUG-011 wholesale `_copy_tree` drift: `files[]` becomes the sole driver of staged bytes;",
        "local Docker targets are first-class in `deployment_architecture.json`; Crosswalk runs at",
        "commit, build, local deploy, and cloud deploy.",
        "",
        "## Five new targets (local)",
        "",
        "| target_id | kind | node_address | cloud mirror |",
        "|-----------|------|--------------|--------------|",
    ]
    for s in specs:
        kind = "docker_local_container" if s.deployable and "docker" in s.target_id else "host_os_process"
        mirror = s.cloud_mirror or "(governance only)"
        lines.append(f"| `{s.target_id}` | {kind} | {s.node_address} | {mirror} |")

    lines.extend([
        "",
        "## Reachability summary",
        "",
        f"Git-tracked source rows analyzed: **{stats['tracked_files']}**",
        "",
        "| disposition | count |",
        "|-------------|-------|",
    ])
    for k, v in sorted(stats["disposition"].items()):
        lines.append(f"| {k} | {v} |")

    lines.extend([
        "",
        "| target | semantic (import walk) | structural (Dockerfile COPY) | declared files[] (mirror) | wholesale huggingface_source_set |",
        "|--------|------------------------|------------------------------|---------------------------|-----------------------------------|",
    ])
    for tid, tstats in stats["targets"].items():
        lines.append(
            f"| `{tid}` | {tstats['semantic']} | {tstats['structural']} | "
            f"{tstats['declared_mirror']} | {tstats['wholesale_mirror']} |"
        )

    lines.extend([
        "",
        "## Proposed build pipeline changes",
        "",
        "1. **Stage from `files[]` only** — Replace `hf_helpers._copy_tree(huggingface_source_set)` with",
        "   `Extractor.copy_referenced_files(record.files)` for HF and local Docker contexts.",
        "2. **Wire `Builder.build_record()`** — On every build/deploy, refresh `content_hash` and optionally",
        "   embed managed dockerfiles; reject if semantic reachability agent finds gaps.",
        "3. **`LocalDeploy` reads manifest** — Replace `BACKEND_CONTAINERS` hard-code with",
        "   `target_docker_local_*` records (container name, Dockerfile path, port from `node_address`).",
        "4. **Crosswalk everywhere** — Add `_deployment_architecture_gate()` to `run_cloud_deploy()` before HF git push.",
        "5. **Workstation target** — `target_host_local_workstation` holds DevOps/brain/hook JSON + secrets inventory;",
        "   not deployed to cloud; Crosswalk ensures hook paths stay present.",
        "6. **Deprecate `huggingface_source_set`** — Keep as audit-only cross-check until bloat hits zero, then remove.",
        "",
        "## Spreadsheet columns",
        "",
        f"- **CSV:** `{csv_path}`",
    ])
    if xlsx_path:
        lines.append(f"- **XLSX:** `{xlsx_path}` (includes `summary` sheet)")
    lines.extend([
        f"- **Proposed files[] drafts:** `{proposed_path}`",
        "",
        "Use `proposed_targets` and `gap_vs_declared` columns to refactor `deployment_architecture.json`.",
        "Use `bloat_vs_files` and `wholesale_copy` to trim `huggingface_source_set` and ops/ subtrees.",
        "",
        "## Regenerate",
        "",
        "```powershell",
        "python architecture/DevOpsBuildDeployAndEnvironmentManagement/target_reachability_audit.py",
        "```",
        "",
    ])
    return "\n".join(lines)


if __name__ == "__main__":
    main()
